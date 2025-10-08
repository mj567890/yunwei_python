"""
Excel处理工具类
"""
import io
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app


class ExcelProcessor:
    """Excel处理器"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self, sheet_name: str = 'Sheet1'):
        """创建工作簿"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        return self
    
    def load_workbook(self, file_path: str, sheet_name: Optional[str] = None):
        """加载工作簿"""
        self.workbook = load_workbook(file_path, data_only=True)
        if sheet_name and sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[sheet_name]
        else:
            self.worksheet = self.workbook.active
        return self
    
    def load_from_stream(self, file_stream, sheet_name: Optional[str] = None):
        """从文件流加载工作簿"""
        self.workbook = load_workbook(file_stream, data_only=True)
        if sheet_name and sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[sheet_name]
        else:
            self.worksheet = self.workbook.active
        return self
    
    def write_headers(self, headers: List[str], row: int = 1):
        """写入表头"""
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=row, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
    
    def write_data(self, data: List[List[Any]], start_row: int = 2):
        """写入数据"""
        for row_index, row_data in enumerate(data, start_row):
            for col_index, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=row_index, column=col_index, value=value)
                # 设置数据样式
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
    
    def auto_adjust_column_width(self):
        """自动调整列宽"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_to_file(self, file_path: str):
        """保存到文件"""
        self.workbook.save(file_path)
    
    def save_to_stream(self) -> io.BytesIO:
        """保存到字节流"""
        stream = io.BytesIO()
        self.workbook.save(stream)
        stream.seek(0)
        return stream
    
    def read_data(self, start_row: int = 2, max_row: Optional[int] = None) -> List[List[Any]]:
        """读取数据"""
        data = []
        max_row = max_row or self.worksheet.max_row
        
        for row in self.worksheet.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
            # 过滤空行
            if any(cell is not None for cell in row):
                data.append(list(row))
        
        return data
    
    def read_headers(self, row: int = 1) -> List[str]:
        """读取表头"""
        headers = []
        for cell in self.worksheet[row]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                break
        return headers
    
    def validate_headers(self, expected_headers: List[str], row: int = 1) -> tuple[bool, List[str]]:
        """验证表头"""
        actual_headers = self.read_headers(row)
        missing_headers = []
        
        for header in expected_headers:
            if header not in actual_headers:
                missing_headers.append(header)
        
        return len(missing_headers) == 0, missing_headers


class AssetExcelProcessor(ExcelProcessor):
    """资产Excel处理器"""
    
    # 资产导出表头
    EXPORT_HEADERS = [
        '资产编码', '资产名称', '品牌', '型号', '资产类别',
        '楼宇', '楼层', '房间', '详细位置',
        '供应商', '采购日期', '采购价格', '采购订单号',
        '保修开始日期', '保修结束日期', '保修期(月)',
        '使用人', '使用部门', '部署日期',
        '资产状态', '状况评级', '序列号', 'MAC地址', 'IP地址', '备注'
    ]
    
    # 资产导入表头映射
    IMPORT_HEADER_MAP = {
        '资产编码': 'asset_code',
        '资产名称': 'name',
        '品牌': 'brand',
        '型号': 'model',
        '资产类别': 'category',
        '楼宇': 'building_name',
        '楼层': 'floor_name',
        '房间': 'room_name',
        '详细位置': 'location_detail',
        '供应商': 'supplier',
        '采购日期': 'purchase_date',
        '采购价格': 'purchase_price',
        '采购订单号': 'purchase_order',
        '保修开始日期': 'warranty_start_date',
        '保修结束日期': 'warranty_end_date',
        '保修期(月)': 'warranty_period',
        '使用人': 'user_name',
        '使用部门': 'user_department',
        '部署日期': 'deploy_date',
        '资产状态': 'status',
        '状况评级': 'condition_rating',
        '序列号': 'serial_number',
        'MAC地址': 'mac_address',
        'IP地址': 'ip_address',
        '备注': 'remark'
    }
    
    def export_assets(self, assets: List[Dict]) -> io.BytesIO:
        """导出资产数据"""
        self.create_workbook('资产列表')
        
        # 写入表头
        self.write_headers(self.EXPORT_HEADERS)
        
        # 准备数据
        data = []
        for asset in assets:
            row = [
                asset.get('asset_code', ''),
                asset.get('name', ''),
                asset.get('brand', ''),
                asset.get('model', ''),
                asset.get('category', ''),
                asset.get('building_name', ''),
                asset.get('floor_name', ''),
                asset.get('room_name', ''),
                asset.get('location_detail', ''),
                asset.get('supplier', ''),
                asset.get('purchase_date', ''),
                asset.get('purchase_price', ''),
                asset.get('purchase_order', ''),
                asset.get('warranty_start_date', ''),
                asset.get('warranty_end_date', ''),
                asset.get('warranty_period', ''),
                asset.get('user_name', ''),
                asset.get('user_department', ''),
                asset.get('deploy_date', ''),
                asset.get('status', ''),
                asset.get('condition_rating', ''),
                asset.get('serial_number', ''),
                asset.get('mac_address', ''),
                asset.get('ip_address', ''),
                asset.get('remark', '')
            ]
            data.append(row)
        
        # 写入数据
        self.write_data(data)
        
        # 自动调整列宽
        self.auto_adjust_column_width()
        
        return self.save_to_stream()
    
    def import_assets(self, file_stream) -> Dict[str, Any]:
        """导入资产数据"""
        try:
            self.load_from_stream(file_stream)
            
            # 验证表头
            expected_headers = list(self.IMPORT_HEADER_MAP.keys())
            is_valid, missing_headers = self.validate_headers(expected_headers)
            
            if not is_valid:
                return {
                    'success': False,
                    'message': f'表头验证失败，缺少字段: {", ".join(missing_headers)}',
                    'data': None
                }
            
            # 读取数据
            headers = self.read_headers()
            data_rows = self.read_data()
            
            if not data_rows:
                return {
                    'success': False,
                    'message': '文件中没有数据',
                    'data': None
                }
            
            # 解析数据
            assets = []
            errors = []
            
            for row_index, row in enumerate(data_rows, 2):  # 从第2行开始
                try:
                    asset_data = {}
                    
                    for col_index, header in enumerate(headers):
                        if col_index < len(row) and header in self.IMPORT_HEADER_MAP:
                            field_name = self.IMPORT_HEADER_MAP[header]
                            value = row[col_index]
                            
                            # 数据类型转换和验证
                            if value is not None:
                                value = str(value).strip()
                                if value:
                                    asset_data[field_name] = self._convert_field_value(field_name, value)
                    
                    # 必填字段验证
                    if not asset_data.get('asset_code'):
                        errors.append(f'第{row_index}行: 资产编码不能为空')
                        continue
                    
                    if not asset_data.get('name'):
                        errors.append(f'第{row_index}行: 资产名称不能为空')
                        continue
                    
                    assets.append(asset_data)
                    
                except Exception as e:
                    errors.append(f'第{row_index}行: 数据解析错误 - {str(e)}')
            
            return {
                'success': True,
                'message': f'解析完成，共{len(assets)}条有效数据',
                'data': {
                    'assets': assets,
                    'errors': errors,
                    'total_rows': len(data_rows),
                    'valid_rows': len(assets),
                    'error_rows': len(errors)
                }
            }
            
        except Exception as e:
            current_app.logger.error(f'导入资产数据失败: {str(e)}')
            return {
                'success': False,
                'message': f'文件解析失败: {str(e)}',
                'data': None
            }
    
    def _convert_field_value(self, field_name: str, value: str):
        """转换字段值"""
        if field_name in ['purchase_price']:
            try:
                return float(value)
            except:
                return None
        
        elif field_name in ['warranty_period']:
            try:
                return int(value)
            except:
                return None
        
        elif field_name in ['purchase_date', 'warranty_start_date', 'warranty_end_date', 'deploy_date']:
            try:
                # 尝试多种日期格式
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日']:
                    try:
                        return datetime.strptime(value, fmt).date()
                    except:
                        continue
                return None
            except:
                return None
        
        else:
            return value
    
    def generate_template(self) -> io.BytesIO:
        """生成导入模板"""
        self.create_workbook('资产导入模板')
        
        # 写入表头
        self.write_headers(list(self.IMPORT_HEADER_MAP.keys()))
        
        # 添加示例数据
        sample_data = [
            [
                'AS20240001', 'Dell服务器01', 'Dell', 'PowerEdge R740', '服务器',
                '总部大楼', '1层', '机房101', '机柜A01',
                'Dell中国', '2024-01-15', '85000', 'PO20240001',
                '2024-01-15', '2027-01-15', '36',
                '张三', 'IT部', '2024-01-20',
                '在用', '优', 'DL001SN2024001', '00:1B:44:11:3A:B7', '192.168.1.100', '核心服务器'
            ]
        ]
        
        self.write_data(sample_data)
        
        # 自动调整列宽
        self.auto_adjust_column_width()
        
        return self.save_to_stream()